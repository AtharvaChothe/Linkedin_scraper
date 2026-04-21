from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests
from bs4 import BeautifulSoup
import time
import random
import re
from urllib.parse import urljoin, urlparse, quote_plus
import json
import logging

app = Flask(__name__)
CORS(app)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# ─── HEADERS ──────────────────────────────────────────────────────────────────

def get_stealth_headers():
    return {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
    }

def get_google_headers():
    return {
        'User-Agent': random.choice([
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        ]),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Referer': 'https://www.google.com/',
    }

def get_random_headers():
    return {
        "User-Agent": random.choice([
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
        ]),
        "Accept-Language": "en-US,en;q=0.9",
        "Connection": "keep-alive"
    }

def human_delay(min_s=1.5, max_s=10):
    time.sleep(random.uniform(min_s, max_s))


def normalize_linkedin_url(url):
    if not url:
        raise ValueError("URL required")
    url = url.strip()
    if not re.match(r'^https?://', url, re.IGNORECASE):
        url = 'https://' + url
    parsed = urlparse(url)
    if not parsed.netloc:
        raise ValueError("Invalid URL")
    clean_path = re.sub(r'(?<!:)//+', '/', parsed.path)
    return f"{parsed.scheme}://{parsed.netloc}{clean_path}".rstrip('/')


def convert_count_suffix(count_str):
    """Convert K/M/B suffixes to full numbers"""
    if not count_str:
        return count_str
    
    original = count_str
    count_str = count_str.upper().strip()
    multiplier = 1
    
    if count_str.endswith('K'):
        multiplier = 1000
        count_str = count_str[:-1]
    elif count_str.endswith('M'):
        multiplier = 1000000
        count_str = count_str[:-1]
    elif count_str.endswith('B'):
        multiplier = 1000000000
        count_str = count_str[:-1]
    
    try:
        # Handle decimal numbers like 1.5K
        if '.' in count_str:
            num = float(count_str)
        else:
            num = int(count_str.replace(',', ''))
        
        total = int(num * multiplier)
        return f"{total:,}"
    except:
        return original


# ─── EMPLOYEE COUNT EXTRACTOR (IMPROVED) ─────────────────────────────────────────────────

def extract_employee_count(soup_main, soup_about, raw_html_main, raw_html_about):
    """
    Extracts employee count with multiple strategies.
    Returns a string like '10,001+ employees' or 'N/A'
    """
    
    # Strategy 1: Look for "X associated members" (most accurate for LinkedIn)
    for html in [raw_html_about, raw_html_main]:
        # Pattern for "X associated members"
        patterns = [
            r'([\d,]+(?:\.\d+)?[KkMmBb]?\+?)\s+associated\s+members?',
            r'([\d,]+(?:\.\d+)?[KkMmBb]?\+?)\s+members?',
            r'([\d,]+(?:\.\d+)?[KkMmBb]?\+?)\s+employees?',
        ]
        for pattern in patterns:
            m = re.search(pattern, html, re.IGNORECASE)
            if m:
                count = m.group(1).strip()
                # Convert K/M/B suffixes
                count = convert_count_suffix(count)
                # Get the full match for context
                full_match = m.group(0).strip()
                return full_match
    
    # Strategy 2: dt/dd pairs with size/employee label
    for soup_target in [soup_about, soup_main]:
        for dt in soup_target.find_all('dt'):
            label = dt.get_text(strip=True).lower()
            if 'size' in label or 'employee' in label or 'headcount' in label:
                dd = dt.find_next_sibling('dd')
                if dd:
                    val = dd.get_text(strip=True)
                    if val and any(c.isdigit() for c in val):
                        return val.strip()
    
    # Strategy 3: Regex on full page text for employee ranges
    for html in [raw_html_about, raw_html_main]:
        soup_tmp = BeautifulSoup(html, 'html.parser')
        full_text = soup_tmp.get_text(separator=' ', strip=True)
        
        # Look for patterns like "10,001+ employees" or "51-200 employees"
        patterns = [
            r'(\d{1,3}(?:,\d{3})*\+?\s*(?:-|to)?\s*\d{0,3}(?:,\d{3})*\+?)\s*employees?',
            r'employees?\s*:\s*(\d{1,3}(?:,\d{3})*\+?)',
            r'staff\s*:\s*(\d{1,3}(?:,\d{3})*\+?)',
            r'headcount\s*:\s*(\d{1,3}(?:,\d{3})*\+?)',
        ]
        for pattern in patterns:
            m = re.search(pattern, full_text, re.IGNORECASE)
            if m:
                return m.group(0).strip()
    
    # Strategy 4: JSON blobs in the page
    for html in [raw_html_about, raw_html_main]:
        json_patterns = [
            r'"staffCount"\s*:\s*"([^"]+)"',
            r'"staffCount"\s*:\s*(\d+)',
            r'"employeeCount"\s*:\s*"([^"]+)"',
            r'"employeeCount"\s*:\s*(\d+)',
            r'"numEmployees"\s*:\s*"([^"]+)"',
            r'"numEmployees"\s*:\s*(\d+)',
            r'"headcount"\s*:\s*"([^"]+)"',
            r'"headcount"\s*:\s*(\d+)',
            r'"size"\s*:\s*"([^"]+)"',
        ]
        for pattern in json_patterns:
            m = re.search(pattern, html, re.IGNORECASE)
            if m:
                count = m.group(1).strip()
                if count.isdigit():
                    return f"{int(count):,} employees"
                return f"{count} employees"
    
    # Strategy 5: JSON-LD numberOfEmployees
    for soup_target in [soup_about, soup_main]:
        for script in soup_target.find_all('script', type='application/ld+json'):
            try:
                ld = json.loads(script.string or "")
                if isinstance(ld, list):
                    for item in ld:
                        if item.get('@type') == 'Organization':
                            count = item.get('numberOfEmployees')
                            if count:
                                if isinstance(count, dict):
                                    val = count.get('value')
                                    if val:
                                        return f"{int(val):,} employees"
                                elif isinstance(count, (int, float)):
                                    return f"{int(count):,} employees"
                                elif isinstance(count, str) and count.isdigit():
                                    return f"{int(count):,} employees"
            except Exception:
                pass
    
    # Strategy 6: Look for LinkedIn size range strings
    for html in [raw_html_about, raw_html_main]:
        # Standard LinkedIn size categories
        size_patterns = [
            r'(Self-employed|1-10|11-50|51-200|201-500|501-1,000|1,001-5,000|5,001-10,000|10,001\+)\s*employees?',
            r'(\d+[+-]\d+|\d+\+|\d+-\d+)\s*employees?',
        ]
        for pattern in size_patterns:
            m = re.search(pattern, html, re.IGNORECASE)
            if m:
                return m.group(0).strip()
    
    # Strategy 7: Extract from meta tags
    for soup_target in [soup_about, soup_main]:
        meta_desc = soup_target.find('meta', {'name': 'description'})
        if meta_desc and meta_desc.get('content'):
            m = re.search(r'(\d{1,3}(?:,\d{3})*\+?)\s*employees?', meta_desc['content'], re.IGNORECASE)
            if m:
                return m.group(0).strip()
    
    return "N/A"


# ─── GOOGLE DORK: find employee LinkedIn URLs without login ───────────────────

def google_dork_employees(company_name, company_slug, max_results=25):
    found = {}
    session = requests.Session()
    session.headers.update(get_google_headers())

    queries = []
    if company_name and company_name != "N/A":
        queries.append(f'site:linkedin.com/in "{company_name}"')
        queries.append(f'site:linkedin.com/in "{company_name}" -jobs')
    if company_slug:
        queries.append(f'site:linkedin.com/in inurl:{company_slug}')

    for query in queries:
        if len(found) >= max_results:
            break
        try:
            search_url = f"https://www.google.com/search?q={quote_plus(query)}&num=20&hl=en&gl=us"
            print(f"   Google dork: {query}")
            resp = session.get(search_url, timeout=12)

            if resp.status_code == 429:
                print("   Google rate-limited — stopping dork")
                break
            if resp.status_code != 200:
                print(f"   Google status {resp.status_code}")
                continue

            soup = BeautifulSoup(resp.text, 'html.parser')

            for a in soup.find_all('a', href=True):
                href = a['href']
                if '/url?q=' in href:
                    href = re.sub(r'^.*?/url\?q=', '', href)
                    href = re.sub(r'&.*$', '', href)

                if not re.search(r'linkedin\.com/in/[A-Za-z0-9_%-]+', href):
                    continue
                if any(s in href for s in ['authwall', 'login', '/in/search', 'discover']):
                    continue

                clean_url = re.sub(r'\?.*$', '', href).rstrip('/')
                if not clean_url.startswith('http'):
                    clean_url = 'https://www.linkedin.com' + clean_url
                if clean_url in found:
                    continue

                name, title = "N/A", "N/A"
                parent = a
                for _ in range(7):
                    parent = getattr(parent, 'parent', None)
                    if parent is None:
                        break
                    card_text = parent.get_text(separator='\n', strip=True)
                    lines = [l.strip() for l in card_text.splitlines() if l.strip()]
                    if lines:
                        first = re.sub(r'\s*\|\s*LinkedIn.*$', '', lines[0], flags=re.IGNORECASE).strip()
                        if ' - ' in first or ' \u2013 ' in first:
                            parts = re.split(r' [-\u2013] ', first, maxsplit=1)
                            name  = parts[0].strip()[:80]
                            title = parts[1].strip()[:120] if len(parts) > 1 else "N/A"
                        elif first and len(first) < 80:
                            name = first
                        if name != "N/A":
                            break

                found[clean_url] = {"name": name, "title": title, "url": clean_url}

            print(f"   Running total: {len(found)} profiles")
            human_delay(2.5, 5.0)

        except Exception as e:
            print(f"   Dork error: {str(e)[:100]}")
            continue

    return list(found.values())[:max_results]


# ─── EMPLOYEE SCRAPER ─────────────────────────────────────────────────────────

def scrape_company_employees_enhanced(company_url, session, company_name=""):
    print(f"   Hunting employees for: {company_url}")
    employees = []
    seen_urls  = set()

    def add_employees_from_html(html_text, limit=20):
        added = 0
        for raw in re.findall(
            r'https?://(?:www\.)?linkedin\.com/in/([A-Za-z0-9_%-]{3,80})', html_text
        ):
            full_url = f"https://www.linkedin.com/in/{raw.rstrip('/?')}"
            if full_url not in seen_urls:
                seen_urls.add(full_url)
                employees.append({"name": "N/A", "title": "N/A", "url": full_url})
                added += 1
                if added >= limit:
                    break

        for raw in re.findall(r'(?:href=["\']|"url"\s*:\s*")/in/([A-Za-z0-9_%-]{3,80})', html_text):
            full_url = f"https://www.linkedin.com/in/{raw.rstrip('/?')}"
            if full_url not in seen_urls:
                seen_urls.add(full_url)
                employees.append({"name": "N/A", "title": "N/A", "url": full_url})
                added += 1
                if added >= limit:
                    break
        return added

    # Technique 1: /people/ page
    people_url = company_url.rstrip('/') + '/people/'
    try:
        resp = session.get(people_url, timeout=10, headers=get_stealth_headers())
        final = resp.url.lower()
        blocked = any(b in final for b in ['authwall', 'login', 'signin', 'checkpoint'])

        if not blocked and resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'html.parser')
            for a in soup.find_all('a', href=True):
                href = a['href']
                if '/in/' not in href:
                    continue
                if any(s in href for s in ['authwall', 'login', '/in/search']):
                    continue
                full = urljoin('https://www.linkedin.com', href)
                clean = re.sub(r'\?.*$', '', full).rstrip('/')
                if clean in seen_urls:
                    continue
                seen_urls.add(clean)
                nm = a.get_text(strip=True)
                employees.append({
                    "name":  nm[:80] if nm and len(nm) < 80 else "N/A",
                    "title": "N/A",
                    "url":   clean,
                })
            add_employees_from_html(resp.text, limit=20)
            print(f"   Technique 1 (/people/): {len(employees)} profiles")
        else:
            print("   Technique 1: auth-walled")
    except Exception as e:
        print(f"   Technique 1 error: {str(e)[:80]}")

    # Technique 2: LinkedIn search endpoint
    if len(employees) < 5:
        slug_m = re.search(r'/company/([^/?#]+)', company_url, re.IGNORECASE)
        slug   = slug_m.group(1) if slug_m else ""
        if slug:
            search_url = (
                f"https://www.linkedin.com/search/results/people/"
                f"?currentCompany=%5B%22{slug}%22%5D&origin=COMPANY_PAGE_CURATION"
            )
            try:
                resp = session.get(search_url, timeout=10, headers=get_stealth_headers())
                final = resp.url.lower()
                blocked = any(b in final for b in ['authwall', 'login', 'signin', 'checkpoint'])
                if not blocked and resp.status_code == 200:
                    added = add_employees_from_html(resp.text, limit=20)
                    print(f"   Technique 2 (search): +{added} profiles")
                else:
                    print("   Technique 2: auth-walled")
            except Exception as e:
                print(f"   Technique 2 error: {str(e)[:80]}")

    # Technique 3: Google dork
    if len(employees) < 5:
        slug_m = re.search(r'/company/([^/?#]+)', company_url, re.IGNORECASE)
        slug   = slug_m.group(1) if slug_m else ""
        print(f"   Technique 3: Google dork for '{company_name or slug}'")
        dork_results = google_dork_employees(
            company_name=company_name,
            company_slug=slug,
            max_results=25,
        )
        for emp in dork_results:
            if emp['url'] not in seen_urls:
                seen_urls.add(emp['url'])
                employees.append(emp)
        print(f"   Technique 3 (Google dork): total now {len(employees)} profiles")

    return employees[:30]


def format_employees_for_display(employees):
    if not employees:
        return "N/A"
    lines = []
    for emp in employees:
        name  = emp.get("name",  "N/A") or "N/A"
        title = emp.get("title", "N/A") or "N/A"
        url   = emp.get("url",   "")    or ""
        label = name if name != "N/A" else ""
        if title != "N/A":
            label = f"{label} | {title}".strip(" |")
        if not label:
            label = url.split("/in/")[-1] if "/in/" in url else "Unknown"
        if url:
            lines.append(f"{label}\n{url}")
        else:
            lines.append(label)
    return "\n\n".join(lines)


# ─── COMPANY SCRAPER ──────────────────────────────────────────────────────────

def scrape_company(url):
    session = requests.Session()
    session.headers.update(get_stealth_headers())
    data = {"LinkedIn URL": url}

    try:
        print(f"Scraping company: {url}")

        # Main page
        resp = session.get(url, timeout=15)
        resp.raise_for_status()
        raw_html_main = resp.text
        soup = BeautifulSoup(raw_html_main, 'html.parser')

        # Company Name
        for sel in [
            'h1[data-test-id="hero-overlay-title"]',
            '.org-page-top-card-summary__title h1',
            '.org-top-card-module__container h1',
            'h1.top-card-layout__title',
            'h1',
        ]:
            el = soup.select_one(sel)
            if el:
                data["Company Name"] = el.get_text(strip=True)
                break

        # Tagline
        for sel in [
            '[data-test-id="hero-overlay-subtitle"]',
            '.org-top-card-module__tagline',
            '.top-card-layout__headline',
        ]:
            el = soup.select_one(sel)
            if el:
                data["Tagline"] = el.get_text(strip=True)
                break

        # Followers
        m = re.search(r'([\d,.KkMm]+)\s*follower[s]?', soup.get_text(), re.IGNORECASE)
        if m:
            data["Followers"] = m.group(1).strip()

        # About page
        about_url = url.rstrip('/') + '/about/'
        resp_about = session.get(about_url, timeout=15)
        raw_html_about = resp_about.text
        soup_about = BeautifulSoup(raw_html_about, 'html.parser')

        for sel in [
            '[data-test-id="about-us__description"]',
            '.org-about-us-organization-description',
            '.org-about-module__description p',
            '.break-words',
        ]:
            el = soup_about.select_one(sel)
            if el:
                raw = el.get_text(strip=True)
                if len(raw) > 50:
                    data["Overview"] = raw[:500] + ("..." if len(raw) > 500 else "")
                    break

        # Structured dt/dd pairs
        dl_items = soup_about.find_all(['dt', 'dd']) or soup.find_all(['dt', 'dd'])
        i = 0
        while i < len(dl_items):
            if dl_items[i].name == 'dt':
                label = dl_items[i].get_text(strip=True).lower()
                if i + 1 < len(dl_items) and dl_items[i + 1].name == 'dd':
                    value = dl_items[i + 1].get_text(strip=True)
                    if 'website'       in label: data["Website"]      = value
                    elif 'phone'       in label or 'telephone' in label or 'contact' in label:
                        data["Phone"] = value
                    elif 'industry'    in label: data["Industry"]     = value
                    elif 'size'        in label or 'employee' in label:
                        data["Company Size"] = value
                    elif 'headquarter' in label or 'hq' in label or 'location' in label:
                        data["Headquarters"] = value
                    elif 'founded'     in label or 'established' in label:
                        data["Founded"]      = value
                    elif 'type'        in label: data["Type"]         = value
                    elif 'specialt'    in label or 'skills' in label:
                        data["Specialties"]  = value
                    i += 2
                else:
                    i += 1
            else:
                i += 1

        # li fallback
        for li in soup_about.find_all('li'):
            tc = li.get_text(strip=True).lower()
            v  = li.get_text(strip=True)
            def _li(key, *keywords):
                if any(k in tc for k in keywords) and not data.get(key):
                    parts = v.split(':')
                    data[key] = parts[-1].strip() if len(parts) > 1 else v
            _li("Website",      'website')
            _li("Phone",        'phone', 'telephone', 'contact')
            _li("Industry",     'industry')
            _li("Company Size", 'size')
            _li("Headquarters", 'headquarter')
            _li("Founded",      'founded', 'established')
            _li("Type",         'type')
            _li("Specialties",  'specialt', 'skills')

        # Phone regex fallback
        if not data.get("Phone"):
            for pat in [
                r'\+\d{1,3}[\s\-]?\d{1,14}',
                r'\(\d{3}\)[\s\-]?\d{3}[\s\-]?\d{4}',
                r'\d{3}[\s\-]?\d{3}[\s\-]?\d{4}',
            ]:
                m = re.search(pat, soup_about.get_text())
                if m:
                    data["Phone"] = m.group(0)
                    break

        # JSON-LD fallback
        for script in soup.find_all('script', type='application/ld+json'):
            try:
                ld = json.loads(script.string)
                if isinstance(ld, list): ld = ld[0]
                if ld.get('@type') == 'Organization':
                    data.setdefault("Website",  ld.get('url',      'N/A'))
                    data.setdefault("Industry", ld.get('industry', 'N/A'))
                    break
            except Exception:
                pass

        # ── EMPLOYEE COUNT (IMPROVED EXTRACTION) ────────────────────────────
        employee_count = extract_employee_count(soup, soup_about, raw_html_main, raw_html_about)
        
        # Use employee count if found, otherwise fall back to company size
        if employee_count != "N/A":
            data["Employee Count"] = employee_count
        elif data.get("Company Size") and data["Company Size"] != "N/A":
            data["Employee Count"] = data["Company Size"]
        else:
            data["Employee Count"] = "N/A"
        
        # Also try to extract from the main page raw HTML if not found
        if data["Employee Count"] == "N/A":
            # Look for "X employees" pattern in main page
            main_match = re.search(r'(\d{1,3}(?:,\d{3})*\+?)\s*employees?', raw_html_main, re.IGNORECASE)
            if main_match:
                data["Employee Count"] = main_match.group(0).strip()
        
        # Additional check for company size from about page
        if data["Employee Count"] == "N/A" and data.get("Company Size") == "N/A":
            # Look for size in about page text
            about_text = soup_about.get_text()
            size_match = re.search(r'(?:Company size|Size)[:\s]+(\d[\d,+\-]+)', about_text, re.IGNORECASE)
            if size_match:
                data["Employee Count"] = size_match.group(1).strip()

        # ── EMPLOYEES (profile links) ────────────────────────────────────────
        company_name = data.get("Company Name", "")
        try:
            emp_list = scrape_company_employees_enhanced(url, session, company_name)
            data["Employees"] = format_employees_for_display(emp_list)
        except Exception as e:
            print(f"Employee scrape error: {e}")
            data["Employees"] = "N/A"

        # Fill missing fields
        for key in [
            "Company Name", "Tagline", "Overview", "Followers",
            "Website", "Phone", "Industry", "Company Size", "Employee Count",
            "Headquarters", "Founded", "Specialties", "Type", "Employees",
        ]:
            data.setdefault(key, "N/A")

    except requests.RequestException as e:
        logging.error(f"Network error scraping company {url}: {str(e)}")
        raise Exception(f"Network error: {str(e)}")
    except Exception as e:
        logging.error(f"Scraping failed for company {url}: {str(e)}")
        raise Exception(f"Scraping failed: {str(e)}")

    human_delay()
    return data


# ─── USER SCRAPER ─────────────────────────────────────────────────────────────

def scrape_user(url):
    session = requests.Session()
    session.headers.update({
        "User-Agent": random.choice([
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        ]),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1", "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none", "Cache-Control": "max-age=0",
    })

    data = {
        "LinkedIn URL": url, "Full Name": "N/A", "Headline": "N/A",
        "Location": "N/A", "About": "N/A", "Current Position": "N/A",
        "Current Company": "N/A", "Education": "N/A",
        "Connections": "N/A", "Profile Picture": "N/A",
    }

    try:
        print(f"Scraping user: {url}")
        resp = session.get(url, timeout=20, allow_redirects=True)
        final_url = resp.url.lower()
        if any(x in final_url for x in ["authwall", "login", "checkpoint", "signin"]):
            print("   Auth wall — extracting from meta only")

        raw_html = resp.text
        soup = BeautifulSoup(raw_html, "html.parser")

        og_title = soup.find("meta", property="og:title")
        og_desc  = soup.find("meta", property="og:description")
        og_img   = soup.find("meta", property="og:image")

        if og_title and og_title.get("content"):
            content = re.sub(r"\s*\|\s*LinkedIn.*$", "", og_title["content"], flags=re.IGNORECASE).strip()
            if " - " in content:
                parts = content.split(" - ", 1)
                data["Full Name"] = parts[0].strip()
                data["Headline"]  = parts[1].strip()
            else:
                data["Full Name"] = content

        if og_desc and og_desc.get("content"):
            desc = og_desc["content"].strip()
            loc_match = re.match(r"^([^·•|]{3,60})\s*[·•|]", desc)
            if loc_match:
                candidate = loc_match.group(1).strip()
                if len(candidate) < 60 and not candidate.endswith("."):
                    data["Location"] = candidate
            if len(desc) > 40:
                data["About"] = desc[:500] + ("..." if len(desc) > 500 else "")

        if og_img and og_img.get("content"):
            data["Profile Picture"] = og_img["content"].strip()

        if data["Full Name"] == "N/A":
            title_tag = soup.find("title")
            if title_tag:
                title_text = re.sub(r"\s*\|\s*LinkedIn.*$", "", title_tag.get_text(strip=True), flags=re.IGNORECASE)
                if " - " in title_text:
                    parts = title_text.split(" - ", 1)
                    data["Full Name"] = parts[0].strip()
                    if data["Headline"] == "N/A":
                        data["Headline"] = parts[1].strip()
                elif title_text:
                    data["Full Name"] = title_text.strip()

        json_patterns = {
            "Full Name":        [r'"firstName"\s*:\s*"([^"]{2,60})".*?"lastName"\s*:\s*"([^"]{2,60})"',
                                 r'"name"\s*:\s*"([A-Z][a-zA-Z\s]{2,60})"'],
            "Headline":         [r'"headline"\s*:\s*"([^"]{3,200})"'],
            "Location":         [r'"locationName"\s*:\s*"([^"]{2,100})"',
                                 r'"geoLocationName"\s*:\s*"([^"]{2,100})"'],
            "About":            [r'"summary"\s*:\s*"([^"]{20,2000})"'],
            "Current Position": [
                r'"currentPosition"\s*:\s*\{[^}]*?"title"\s*:\s*"([^\"]{3,200})"',
                r'"title"\s*:\s*"([^\"]{3,200})"',
                r'"jobTitle"\s*:\s*"([^\"]{3,200})"',
            ],
            "Current Company":  [
                r'"currentCompany"\s*:\s*\{[^}]*?"name"\s*:\s*"([^\"]{2,100})"',
                r'"companyName"\s*:\s*"([^\"]{2,100})"',
                r'"worksFor"\s*:\s*\{[^}]*?"name"\s*:\s*"([^\"]{2,100})"',
            ],
            "Connections":      [r'"numConnections"\s*:\s*(\d{1,6})'],
        }
        for field, patterns in json_patterns.items():
            if data[field] != "N/A":
                continue
            for pattern in patterns:
                match = re.search(pattern, raw_html)
                if match:
                    if match.lastindex and match.lastindex >= 2:
                        value = f"{match.group(1).strip()} {match.group(2).strip()}"
                    else:
                        value = match.group(1).strip()
                    value = value.encode("utf-8").decode("unicode_escape", errors="replace")
                    value = re.sub(r"\\n", " ", value).strip()
                    if field == "About":
                        value = value[:500] + ("..." if len(value) > 500 else "")
                    data[field] = value
                    break

        for script in soup.find_all("script", type="application/ld+json"):
            try:
                ld = json.loads(script.string or "")
                if isinstance(ld, list):
                    ld = next((x for x in ld if x.get("@type") == "Person"), {})
                if ld.get("@type") != "Person":
                    continue
                if data["Full Name"]       == "N/A": data["Full Name"]       = ld.get("name", "N/A")
                if data["Headline"]        == "N/A": data["Headline"]        = ld.get("jobTitle", "N/A")
                if data["Current Company"] == "N/A":
                    org = ld.get("worksFor", [{}])
                    if isinstance(org, list): org = org[0] if org else {}
                    data["Current Company"] = org.get("name", "N/A")
                if data["Location"] == "N/A":
                    addr = ld.get("address", {})
                    data["Location"] = addr.get("addressLocality", "N/A") if isinstance(addr, dict) else "N/A"
                break
            except Exception:
                pass

        found = sum(1 for v in data.values() if v not in ("N/A", url))
        print(f"   {found}/{len(data)-1} fields extracted")

    except requests.RequestException as e:
        logging.error(f"Network error scraping user {url}: {e}")
        raise Exception(f"Network error: {e}")
    except Exception as e:
        logging.error(f"Scraping error for user {url}: {e}")
        raise Exception(f"Scraping error: {e}")

    human_delay()
    return data


# ─── EXCEL GENERATOR ──────────────────────────────────────────────────────────

def generate_excel_bytes(data_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LinkedIn Data"

    if data_list and "Company Name" in data_list[0]:
        headers = [
            "Company Name", "Tagline", "Overview", "Followers",
            "Website", "Phone", "Industry", "Company Size", "Employee Count",
            "Employees", "Headquarters", "Founded", "Specialties", "Type", "LinkedIn URL",
        ]
    elif data_list and "Full Name" in data_list[0]:
        headers = [
            "Full Name", "Headline", "Location", "About",
            "Current Position", "Current Company", "Education",
            "Connections", "Profile Picture", "LinkedIn URL",
        ]
    else:
        all_keys = set()
        for item in data_list:
            all_keys.update(item.keys())
        headers = sorted(all_keys)

    header_fill = PatternFill("solid", fgColor="0A66C2")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    col_widths  = {
        "Overview": 70, "Specialties": 60, "Tagline": 50, "About": 70,
        "Employees": 80, "LinkedIn URL": 45, "Profile Picture": 50,
        "Employee Count": 25,
    }

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = col_widths.get(header, 25)

    for row_idx, record in enumerate(data_list, start=2):
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key, "N/A"))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ─── API ROUTES ───────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_file("index.html")


@app.route("/scrape-profile", methods=["POST"])
def scrape_profile():
    body      = request.get_json(silent=True) or {}
    input_url = (body.get("url") or "").strip()
    if not input_url:
        return jsonify({"error": "URL required"}), 400
    try:
        url = normalize_linkedin_url(input_url)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    lower      = url.lower()
    is_company = bool(re.search(r"linkedin\.com/(company|school|showcase)/", lower))
    is_user    = bool(re.search(r"linkedin\.com/(in|pub)/", lower))
    if not (is_company or is_user):
        return jsonify({"error": "Provide a LinkedIn /company/ or /in/ URL"}), 400

    try:
        result = scrape_company(url) if is_company else scrape_user(url)
        return jsonify([result])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/scrape-bulk", methods=["POST"])
def scrape_bulk():
    body = request.get_json(silent=True) or {}
    urls = body.get("urls")
    if not isinstance(urls, list) or not urls:
        return jsonify({"error": "A list of URLs is required"}), 400

    results, errors = [], []
    for raw_url in urls:
        try:
            if not raw_url or not str(raw_url).strip():
                raise ValueError("Empty URL")
            clean_url  = normalize_linkedin_url(str(raw_url))
            lower_url  = clean_url.lower()
            is_company = bool(re.search(r"linkedin\.com/(company|school|showcase)/", lower_url))
            is_user    = bool(re.search(r"linkedin\.com/(in|pub)/", lower_url))
            if not (is_company or is_user):
                errors.append({"url": raw_url, "error": "Must be /company/ or /in/ URL"})
                continue
            results.append(scrape_company(clean_url) if is_company else scrape_user(clean_url))
        except Exception as e:
            errors.append({"url": raw_url, "error": str(e)})

    return jsonify({"results": results, "errors": errors})


@app.route("/download-excel", methods=["POST"])
def download_excel():
    data = request.get_json(silent=True)
    if isinstance(data, dict):
        data_list = data.get("results", [])
    elif isinstance(data, list):
        data_list = data
    else:
        return jsonify({"error": "Invalid data format"}), 400

    if not data_list:
        return jsonify({"error": "No data to export"}), 400

    bio = generate_excel_bytes(data_list)
    return send_file(
        bio,
        as_attachment=True,
        download_name="linkedin_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/upload-urls", methods=["POST"])
def upload_urls():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    if file:
        try:
            filename = file.filename.lower()
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                # Handle Excel file
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                urls = []
                for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):  # Start from row 2, first column
                    url = row[0]
                    if url and str(url).strip():
                        urls.append(str(url).strip())
            else:
                # Handle text/CSV file
                content = file.read().decode('utf-8')
                urls = [line.strip() for line in content.splitlines() if line.strip()]
            if not urls:
                return jsonify({"error": "No URLs found in file"}), 400
        except Exception as e:
            return jsonify({"error": f"Error reading file: {str(e)}"}), 400

    results, errors = [], []
    for raw_url in urls:
        try:
            if not raw_url or not str(raw_url).strip():
                raise ValueError("Empty URL")
            clean_url  = normalize_linkedin_url(str(raw_url))
            lower_url  = clean_url.lower()
            is_company = bool(re.search(r"linkedin\.com/(company|school|showcase)/", lower_url))
            is_user    = bool(re.search(r"linkedin\.com/(in|pub)/", lower_url))
            if not (is_company or is_user):
                errors.append({"url": raw_url, "error": "Must be /company/ or /in/ URL"})
                continue
            results.append(scrape_company(clean_url) if is_company else scrape_user(clean_url))
        except Exception as e:
            errors.append({"url": raw_url, "error": str(e)})

    return jsonify({"results": results, "errors": errors})


if __name__ == "__main__":
    logging.info("LinkedIn Scraper API running at http://localhost:3000")
    logging.info("Employee Count: Extracted from page text, dt/dd pairs, JSON blobs, JSON-LD")
    logging.info("Use responsibly — public data only.")

    import sys
    if len(sys.argv) > 1 and "linkedin.com" in sys.argv[1]:
        url = sys.argv[1]
        try:
            result = scrape_company(url) if "/company/" in url else scrape_user(url)
            print("\n=== SCRAPED DATA ===")
            for k, v in result.items():
                # Truncate long values for display
                display_v = str(v)[:200] + "..." if len(str(v)) > 200 else str(v)
                print(f"  {k}: {display_v}")
        except Exception as e:
            print(f"Error: {e}")

    app.run(debug=True, port=3000, host='0.0.0.0')
